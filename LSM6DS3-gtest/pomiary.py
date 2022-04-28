from cmath import sqrt
import RPi.GPIO as GPIO
import time
import sys
import smbus2
import numpy as np
from openpyxl import Workbook

bus = smbus2.SMBus(1)
adr = 0x6a

wb = Workbook()
sheet = wb.active

def read():
    x = np.int16(bus.read_word_data(adr, 0x28))
    y = np.int16(bus.read_word_data(adr, 0x2A))
    z = np.int16(bus.read_word_data(adr, 0x2C))

    vector = np.array([x,y,z],dtype='int16')

    return vector 

bus.write_byte_data(adr, 0x01, 0b00000000)  # FUNC_CFG_ACCESS (01h)
bus.write_byte_data(adr, 0x04, 0b00000100)  # SENSOR_SYNC_TIME_FRAME (04h)
bus.write_byte_data(adr, 0x10, 0b01011010)  # CTRL1_XL (10h) Acc f [Hz] 208
bus.write_byte_data(adr, 0x11, 0b01010110)  # CTRL2_G (11h)

j = 1
n = 1000
s = '1'
while s != '0':

    suma = 0
    for i in range(n):
        time.sleep(0.01)
        wektor = read()
        suma += wektor / n

    sheet.cell(row=j, column=1).value = suma[0]
    sheet.cell(row=j, column=2).value = suma[1]
    sheet.cell(row=j, column=3).value = suma[2]
    j += 1
    print(suma)
    s = input("Napisz: ")
wb.save("pomiary.xlsx")